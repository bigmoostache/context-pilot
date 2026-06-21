#!/usr/bin/env python3
"""Parse XLSX files from ./drop into YAML in ./parsed/"""

import os
import yaml
import datetime
from openpyxl import load_workbook

DROP = "./drop"
PARSED = "./parsed"

XLSX_FILES = [
    "1.1.3 DM Balance générale FY21.xlsx",
    "1.1.3 DM Balance générale FY22.xlsx",
]

def slugify(name):
    """Create a filesystem-safe slug from filename (without extension)."""
    base = os.path.splitext(name)[0]
    return base.lower().replace(" ", "-").replace("é", "e").replace("è", "e")

def parse_xlsx(filepath):
    """Extract metadata and content from an XLSX file."""
    stat = os.stat(filepath)
    wb = load_workbook(filepath, data_only=True)

    metadata = {
        "source_file": os.path.basename(filepath),
        "file_size_bytes": stat.st_size,
        "modified_utc": datetime.datetime.utcfromtimestamp(stat.st_mtime).isoformat(),
        "format": "xlsx",
        "sheet_count": len(wb.sheetnames),
        "sheet_names": wb.sheetnames,
    }

    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = {
            "dimensions": ws.dimensions,
            "min_row": ws.min_row,
            "max_row": ws.max_row,
            "min_col": ws.min_column,
            "max_col": ws.max_column,
            "row_count": ws.max_row - ws.min_row + 1 if ws.max_row else 0,
            "col_count": ws.max_column - ws.min_column + 1 if ws.max_column else 0,
            "merged_cells": [str(m) for m in ws.merged_cells.ranges] if ws.merged_cells else [],
            "rows": [],
        }

        for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row,
                                 min_col=ws.min_column, max_col=ws.max_column):
            row_data = {}
            for cell in row:
                if cell.value is not None:
                    val = cell.value
                    # Convert datetime objects to ISO strings for YAML
                    if isinstance(val, datetime.datetime):
                        val = val.isoformat()
                    elif isinstance(val, datetime.date):
                        val = val.isoformat()
                    row_data[cell.coordinate] = {
                        "value": val,
                        "row": cell.row,
                        "col": cell.column,
                        "col_letter": cell.column_letter,
                    }
                    if cell.number_format and cell.number_format != "General":
                        row_data[cell.coordinate]["number_format"] = cell.number_format
            if row_data:
                sheet_data["rows"].append(row_data)

        sheets[sheet_name] = sheet_data

    return {"metadata": metadata, "sheets": sheets}


def main():
    os.makedirs(PARSED, exist_ok=True)

    for filename in XLSX_FILES:
        filepath = os.path.join(DROP, filename)
        if not os.path.exists(filepath):
            print(f"  SKIP {filename} (not found)")
            continue

        print(f"  Parsing {filename}...")
        result = parse_xlsx(filepath)
        slug = slugify(filename)
        outdir = os.path.join(PARSED, slug)
        os.makedirs(outdir, exist_ok=True)

        outpath = os.path.join(outdir, "parsed.yaml")
        with open(outpath, "w", encoding="utf-8") as f:
            yaml.dump(result, f, default_flow_style=False, allow_unicode=True, sort_keys=False, width=200)

        print(f"  -> {outpath} ({len(result['sheets'])} sheets)")


if __name__ == "__main__":
    main()
